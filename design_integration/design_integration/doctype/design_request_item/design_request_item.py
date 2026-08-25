import frappe
from frappe import _
from frappe.model.document import Document
from frappe.utils import now_datetime
from frappe.utils import getdate
from frappe.utils import now_datetime
from frappe.utils import flt
from frappe.utils import cint
import frappe.model.naming
import csv
import io
import os
import re
import tempfile
from urllib.parse import quote, unquote

import requests
from openpyxl import load_workbook

GENERATED_BARCODE_TYPE = "CODE-39"

class DesignRequestItem(Document):
    def autoname(self):
        """Generate name for Design Request Item"""
        if not self.name:
            # Get the next number in the series
            last_item = frappe.get_all(
                "Design Request Item",
                fields=["name"],
                order_by="name desc",
                limit=1
            )
            
            if last_item:
                try:
                    last_number = int(last_item[0].name.split('-')[-1])
                    next_number = last_number + 1
                except:
                    next_number = 1
            else:
                next_number = 1
            
            self.name = f"DES-IT-{next_number:06d}"
    
    def validate(self):
        """Validate Design Request Item"""
        self.validate_item()
        self.update_current_stage()
        self.validate_terminal_status_requirements()
        self.create_work_order()
        self.validate_revision_reason()
    
    def on_update(self):
        """Handle updates"""
        self.handle_approval_status_change()
        self.log_stage_transition()
        self.handle_field_dependencies()
    
    def validate_item(self):
        """Validate and populate item details"""
        if self.item_code:
            try:
                item = frappe.get_doc("Item", self.item_code)
                self.item_name = item.item_name
                if not self.description:
                    self.description = item.description or ""
            except:
                frappe.throw(_("Item {0} not found").format(self.item_code))
    
    def update_current_stage(self):
        """Update current stage based on design status"""
        self.current_stage = self.design_status
    
    def handle_approval_status_change(self):
        """Handle approval status changes"""
        if self.has_value_changed("approval_status"):
            # If a revision is being requested explicitly
            if self.approval_status == "Revised":
                # Mark revision flag; keep current design_status unchanged
                self.revision_requested = 1
                self.approval_date = now_datetime()
                # Log revision request in stage transition log
                try:
                    self.append("stage_transition_log", {
                        "stage": "revision",
                        "from_status": self.get_doc_before_save().design_status if self.get_doc_before_save() else self.design_status,
                        "to_status": self.design_status,
                        "transition_date": now_datetime(),
                        "transitioned_by": frappe.session.user,
                        "remarks": f"Revision requested: {getattr(self, 'revision_reason', '') or ''}"
                    })
                except Exception:
                    pass
                return
            
            if self.approval_status == "Approved":
                self.approval_date = now_datetime()
                # If there is an active revision request, only Planning User or System Manager can approve
                if getattr(self, "revision_requested", 0):
                    user_roles = set(frappe.get_roles())
                    allowed_roles = {"Planning User", "System Manager"}
                    if not (user_roles & allowed_roles):
                        frappe.throw(_("Only Planning User or System Manager can approve a revision request."))
                    # Revision approved: send item back to Modelling and increment count
                    self.design_status = "Modelling"
                    try:
                        self.revision_count = (self.revision_count or 0) + 1
                    except Exception:
                        self.revision_count = 1
                    self.revision_requested = 0
                else:
                    # Normal approval flow
                    self.design_status = "Design"
            elif self.approval_status == "Rejected":
                # Send back to Approval Drawing; no On Hold state in design_status
                self.design_status = "Approval Drawing"
                self.approval_date = now_datetime()
            elif self.approval_status == "On Hold":
                # Do not change design_status; only approval_status reflects hold
                self.approval_date = now_datetime()
    
    def log_stage_transition(self):
        """Log stage transitions (store as child rows, not raw dicts)"""
        if self.is_new():
            # don't log on first insert
            return
        if self.has_value_changed("design_status"):
            # set timing fields for Gantt
            if not self.start_date and self.design_status and self.design_status != "Pending":
                self.start_date = now_datetime()
            if self.design_status == "Completed" and not self.completion_date:
                self.completion_date = now_datetime()
            self.append("stage_transition_log", {
                "stage": "design_status",
                "from_status": self.get_doc_before_save().design_status if self.get_doc_before_save() else "",
                "to_status": self.design_status,
                "transition_date": now_datetime(),
                "transitioned_by": frappe.session.user,
                "remarks": f"Status changed to {self.design_status}"
            })
    
    def handle_field_dependencies(self):
        """Handle automatic field updates based on dependencies"""
        # Handle new_item_code changes
        if self.has_value_changed("new_item_code") and self.new_item_code:
            self.sku_generated = 1
            self.item_created = 1
            
            # Fetch item name from the selected item
            try:
                item_doc = frappe.get_doc("Item", self.new_item_code)
                self.new_item_name = item_doc.item_name
            except:
                self.new_item_name = ""
            
            frappe.msgprint(_("SKU Generated and Item Created automatically set to Yes."))
        
        # Handle bom_name changes
        if self.has_value_changed("bom_name") and self.bom_name:
            self.bom_created = 1
            frappe.msgprint(_("BOM Created automatically set to Yes."))
        
        # Handle nesting completion
        if self.design_status == "Nesting":
            self.nesting_completed = 1
        
    def create_work_order(self):
        if frappe.db.exists("Work Order", {"design_request_item" : self.name}):
            return
        if self.design_status == "Completed" and self.bom_name:
            variant_of = None
            if variant_of := frappe.db.get_value("Item",self.new_item_code ,"variant_of"):
                variant_of = variant_of
            from erpnext.manufacturing.doctype.work_order.work_order import make_work_order
            wo_doc = make_work_order(
                self.bom_name,
                self.new_item_code,
                self.qty or 1,
                variant_items = variant_of,
                use_multi_level_bom=1
            )
            wo_doc.design_request_item = self.name
            if self.design_request:
                wo_doc.sales_order = frappe.db.get_value("Design Request", self.design_request, "sales_order")
            wo_doc.save(ignore_permissions=True)

    def validate_revision_reason(self):
        if self.revision_requested and not self.revision_reason:
            frappe.throw(
                "Revision Reason is mandatory when Revision Requested is checked."
            )

    def validate_terminal_status_requirements(self):
        if self.design_status not in ("Completed", "Cancelled"):
            return

        missing = []
        if not (self.sku_generated and self.item_created and self.new_item_code):
            missing.append(_("SKU"))
        if not (self.bom_created and self.bom_name):
            missing.append(_("BOM"))

        if missing:
            frappe.throw(_("Cannot mark {0} until {1} is added.").format(self.design_status, ", ".join(missing)))
            
@frappe.whitelist()
def update_design_status(docname, new_status):
    """Update design status from list view"""
    try:
        doc = frappe.get_doc("Design Request Item", docname)
        doc.design_status = new_status
        doc.save()
        return {"success": True}
    except Exception as e:
        frappe.log_error(f"Error updating design status: {str(e)}")
        return {"success": False, "error": str(e)}




def update_approval_status(docname, new_status):
    """Update approval status from list view"""
    try:
        doc = frappe.get_doc("Design Request Item", docname)
        doc.approval_status = new_status
        doc.save()
        return {"success": True}
    except Exception as e:
        frappe.log_error(f"Error updating approval status: {str(e)}")
        return {"success": False, "error": str(e)} 

@frappe.whitelist()
def get_version_meta_data():
    return frappe.get_meta("Design Version")


@frappe.whitelist()
def get_version_list(design_request_item):
    return frappe.get_all("Design Version", filters={"design_request_item" : design_request_item}, fields=[
        "name", "posting_date", "version_tag", "new_version_file", "description"
    ])



@frappe.whitelist()
def delete_version(version_name, design_request_item):
    """Delete a design version"""
    try:
        # Check permissions
        if not frappe.has_permission("Design Version", "delete"):
            frappe.throw(_("You don't have permission to delete versions"))
        
        # Get the version
        version = frappe.get_doc("Design Version", version_name)
        
        # Store reference before deleting
        docname = version.name
        
        # Delete the document
        frappe.delete_doc("Design Version", version_name, ignore_permissions=False)
        
        frappe.db.commit()
        
        return {
            "success": True,
            "message": _("Version deleted successfully"),
            "deleted_docname": docname
        }
        
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Design Version Deletion Error")
        frappe.throw(_("Error deleting version: {0}").format(str(e)))

@frappe.whitelist()
def check_version_tab(version_tag, name):
    return frappe.db.exists("Design Version", {"version_tag" : version_tag, "design_request_item" : name})

@frappe.whitelist()
def get_next_version_tag(design_request_item):
    doc = frappe.get_doc("Design Request Item", design_request_item)

    base_version = f"V{doc.revision_count}"

    versions = frappe.get_all(
        "Design Version",
        filters={"design_request_item": design_request_item},
        pluck="version_tag"
    )

    # sub version of base version
    suffixes = []
    for version in versions:
        if version == base_version:
            suffixes.append(0)
        elif version.startswith(f"{base_version}-"):
            try:
                suffixes.append(int(version.split("-")[-1]))
            except ValueError:
                pass
    if not suffixes:
        return base_version
    else:
        return f"{base_version}-{max(suffixes) + 1}"


HEADER_ALIASES = {
    "part_no": {"part no", "part number", "partno", "item code", "drawing no", "drawing number"},
    "erp_item_code": {"erp item", "erp item code", "existing item", "existing item code"},
    "part_name": {"part name", "asm part name", "assembly name", "item name"},
    "part_description": {"part description", "description name"},
    "qty": {"qty", "quantity", "qty.", "required qty"},
    "uom": {"uom", "unit", "units"},
    "row_type": {"type", "part type", "asm part", "category", "description"},
    "material": {"material"},
    "bounding_box_length": {"bounding box length", "length"},
    "bounding_box_width": {"bounding box width", "width"},
    "sheet_metal_thickness": {"sheet metal thickness", "thickness", "sheet thickness"},
    "mass": {"mass", "weight"},
    "gross_weight": {"gross weight", "gross wt", "gross weight."},
}

SHEET_HINTS = ("SUB BOM", "BOM IMPORT", "BOM_IMPORT", "BOM")
SUB_ASSEMBLY_RE = re.compile(r"sub\s*[- ]?\s*ass(?:y|embly)", re.IGNORECASE)
MAIN_ASSEMBLY_RE = re.compile(r"main\s*[- ]?\s*ass(?:y|embly)", re.IGNORECASE)


@frappe.whitelist()
def generate_bom_from_design_sheet(design_request_item: str):
    """Create child Items, child BOMs and the FG BOM from the imported design workbook."""
    result = {
        "fg_item": None,
        "fg_bom": None,
        "items_created": [],
        "items_reused": [],
        "boms_created": [],
        "boms_reused": [],
        "generated_sku_barcodes": [],
        "warnings": [],
    }

    design_item = frappe.get_doc("Design Request Item", design_request_item)
    _set_bom_import_state(design_item.name, "Processing", "BOM import started.")

    try:
        _validate_import_permissions(design_item)
        fg_item_code = _get_finished_good_item_code(design_item)
        workbook_path = _resolve_bom_workbook(design_item)
        parsed = _parse_bom_workbook(workbook_path, fg_item_code)
        result["fg_item"] = parsed["fg_item_code"]

        _validate_bom_structure(design_item, parsed)
        source_to_item, item_summary = _resolve_or_create_items(design_item, parsed)
        result["items_created"].extend(item_summary["created"])
        result["items_reused"].extend(item_summary["reused"])
        result["generated_sku_barcodes"].extend(item_summary.get("barcodes", []))

        assembly_boms, bom_summary = _create_bom_hierarchy(design_item, parsed, source_to_item)
        result["boms_created"].extend(bom_summary["created"])
        result["boms_reused"].extend(bom_summary["reused"])

        final_fg_bom = assembly_boms["__fg__"]
        final_bom_doc = frappe.get_doc("BOM", final_fg_bom)
        if final_bom_doc.item != fg_item_code or final_bom_doc.docstatus != 1:
            frappe.throw(_("Generated final BOM is not a submitted BOM for {0}.").format(fg_item_code))

        design_item.db_set("bom_name", final_fg_bom, update_modified=False)
        design_item.db_set("bom_created", 1, update_modified=False)
        _set_generated_sku_barcode_log(design_item.name, result["generated_sku_barcodes"])
        _attach_generated_sku_barcode_sheet(design_item.name, result["generated_sku_barcodes"])
        _set_bom_import_state(
            design_item.name,
            "Completed",
            "Created FG BOM {0}. Created {1} Items and {2} BOMs. Reused {3} Items and {4} BOMs.".format(
                final_fg_bom,
                len(result["items_created"]),
                len(result["boms_created"]),
                len(result["items_reused"]),
                len(result["boms_reused"]),
            ),
        )
        result["fg_bom"] = final_fg_bom
        return result
    except Exception:
        frappe.log_error(frappe.get_traceback(), "Design BOM Import Failed")
        _set_bom_import_state(design_item.name, "Failed", frappe.get_traceback())
        frappe.throw(_("BOM import failed. Check BOM Import Log or Error Log for details."))


def _set_bom_import_state(design_item_name, status, log):
    values = {}
    meta = frappe.get_meta("Design Request Item")
    if meta.has_field("custom_bom_import_status"):
        values["custom_bom_import_status"] = status
    if meta.has_field("custom_bom_import_log"):
        values["custom_bom_import_log"] = log[-60000:] if log else ""
    if values:
        frappe.db.set_value("Design Request Item", design_item_name, values, update_modified=False)


def _set_generated_sku_barcode_log(design_item_name, barcode_rows):
    if not frappe.get_meta("Design Request Item").has_field("custom_generated_sku_barcodes"):
        return

    if not barcode_rows:
        value = ""
    else:
        value = "\n".join(
            "{item_code}\t{barcode}\t{barcode_type}\t{item_name}".format(
                item_code=row.get("item_code"),
                barcode=row.get("barcode"),
                barcode_type=row.get("barcode_type"),
                item_name=row.get("item_name"),
            )
            for row in barcode_rows
        )
    frappe.db.set_value(
        "Design Request Item",
        design_item_name,
        "custom_generated_sku_barcodes",
        value,
        update_modified=False,
    )


def _attach_generated_sku_barcode_sheet(design_item_name, barcode_rows):
    meta = frappe.get_meta("Design Request Item")
    if not meta.has_field("custom_generated_sku_barcode_sheet"):
        return

    file_url = ""
    if barcode_rows:
        file_doc = frappe.get_doc(
            {
                "doctype": "File",
                "file_name": f"{design_item_name}-generated-sku-barcodes.csv",
                "attached_to_doctype": "Design Request Item",
                "attached_to_name": design_item_name,
                "is_private": 1,
                "content": _format_generated_sku_barcode_csv(barcode_rows),
            }
        )
        file_doc.insert(ignore_permissions=True)
        file_url = file_doc.file_url

    frappe.db.set_value(
        "Design Request Item",
        design_item_name,
        "custom_generated_sku_barcode_sheet",
        file_url,
        update_modified=False,
    )


def _format_generated_sku_barcode_csv(barcode_rows):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Item Code", "Barcode", "Barcode Type", "Item Name"])
    for row in barcode_rows:
        writer.writerow([row.get("item_code"), row.get("barcode"), row.get("barcode_type"), row.get("item_name")])
    return output.getvalue()


def _validate_import_permissions(design_item):
    if not frappe.has_permission("Design Request Item", "write", doc=design_item):
        frappe.throw(_("You need Write permission on Design Request Item."))
    if not frappe.has_permission("BOM", "create"):
        frappe.throw(_("You need Create permission on BOM."))
    if not frappe.has_permission("BOM", "submit"):
        frappe.throw(_("You need Submit permission on BOM."))


def _get_finished_good_item_code(design_item):
    return design_item.new_item_code or design_item.item_code


def _resolve_bom_workbook(design_item):
    file_value = (
        getattr(design_item, "custom_bom_for_import", None)
        or getattr(design_item, "custom_bom_importer", None)
        or ""
    ).strip()
    if not file_value:
        frappe.throw(_("Attach an Excel file or add a public Google Sheet link in BOM For Import."))

    if "docs.google.com/spreadsheets" in file_value:
        return _download_google_sheet(file_value)

    file_url = _clean_text(file_value)
    file_doc_name = _find_bom_file_doc(file_url)
    if file_doc_name:
        file_doc = frappe.get_doc("File", file_doc_name)
        file_path = file_doc.get_full_path()
    else:
        file_path = _get_bom_file_path_candidates(file_url)[0]

    file_path_candidates = _get_bom_file_path_candidates(file_url)
    if not os.path.exists(file_path):
        for candidate in file_path_candidates[1:]:
            if os.path.exists(candidate):
                file_path = candidate
                break
        else:
            file_path = _find_bom_file_by_name(file_url)
            if not file_path:
                frappe.throw(
                    _("BOM import file was not found: {0}<br>Checked paths:<br>{1}").format(
                        file_value,
                        "<br>".join(file_path_candidates),
                    )
                )
    if os.path.splitext(file_path)[1].lower() not in (".xlsx", ".xlsm", ".csv"):
        frappe.throw(_("Only .xlsx, .xlsm and .csv files are supported."))
    return file_path


def _find_bom_file_doc(file_url):
    for candidate in _get_bom_file_url_candidates(file_url):
        file_doc_name = frappe.db.get_value("File", {"file_url": candidate}, "name")
        if file_doc_name:
            return file_doc_name

    file_name = os.path.basename(unquote(file_url))
    if file_name:
        return frappe.db.get_value("File", {"file_name": file_name}, "name")
    return None


def _get_bom_file_url_candidates(file_url):
    decoded_url = unquote(file_url)
    dirname, basename = os.path.split(decoded_url)
    encoded_url = os.path.join(dirname, quote(basename)) if basename else decoded_url
    candidates = []
    for candidate in (file_url, decoded_url, encoded_url):
        if candidate and candidate not in candidates:
            candidates.append(candidate)
    return candidates


def _get_bom_file_path_candidates(file_url):
    candidates = []
    for candidate in _get_bom_file_url_candidates(file_url):
        if candidate.startswith(("/files/", "/private/")):
            path = frappe.get_site_path(candidate.lstrip("/"))
        else:
            path = candidate
        if path and path not in candidates:
            candidates.append(path)
    return candidates or [file_url]


def _find_bom_file_by_name(file_url):
    target = _normalize_file_lookup_name(os.path.basename(unquote(file_url)))
    if not target:
        return None

    for folder in ("private/files", "public/files"):
        folder_path = frappe.get_site_path(folder)
        if not os.path.isdir(folder_path):
            continue
        for file_name in os.listdir(folder_path):
            if _normalize_file_lookup_name(file_name) == target:
                return os.path.join(folder_path, file_name)
    return None


def _normalize_file_lookup_name(file_name):
    return re.sub(r"\s+", " ", unquote(_clean_text(file_name))).strip().lower()


def _download_google_sheet(url):
    match = re.search(r"/spreadsheets/d/([^/]+)", url)
    if not match:
        frappe.throw(_("Invalid Google Sheets URL."))

    export_url = f"https://docs.google.com/spreadsheets/d/{match.group(1)}/export?format=xlsx"
    response = requests.get(export_url, timeout=30)
    content_type = response.headers.get("content-type", "")
    if response.status_code != 200 or "text/html" in content_type:
        frappe.throw(_("The Google Sheet cannot be accessed. Make the sheet accessible through the link or upload the Excel file directly."))

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp.write(response.content)
    tmp.close()
    return tmp.name


def _parse_bom_workbook(path, fg_item_code):
    if os.path.splitext(path)[1].lower() == ".csv":
        with open(path, newline="", encoding="utf-8-sig") as csv_file:
            parsed = _parse_table_rows(list(csv.reader(csv_file)))
        if not parsed["assemblies"]:
            frappe.throw(_("No SUB ASSY sections were found in the CSV file."))
        parsed["fg_item_code"] = fg_item_code
        parsed["selected_sheet"] = os.path.basename(path)
        return parsed

    workbook = load_workbook(path, data_only=True, read_only=False, keep_links=False)
    candidates = []
    for sheet in workbook.worksheets:
        if sheet.sheet_state != "visible":
            continue
        parsed = _parse_sheet(sheet)
        if parsed["assemblies"]:
            score = len(parsed["rows"])
            if any(hint in sheet.title.upper() for hint in SHEET_HINTS):
                score += 1000
            candidates.append((score, sheet.title, parsed))

    if not candidates:
        frappe.throw(_("No SUB ASSY sections were found in the workbook."))

    candidates.sort(key=lambda row: row[0], reverse=True)
    selected = candidates[0][2]
    selected["fg_item_code"] = fg_item_code
    selected["selected_sheet"] = candidates[0][1]
    return selected


def _parse_sheet(sheet):
    table = [[cell.value for cell in row] for row in sheet.iter_rows(values_only=False)]
    return _parse_table_rows(table)


def _parse_table_rows(table):
    header_row, columns = _detect_header(table)
    if not header_row:
        return {"assemblies": [], "main_components": [], "rows": []}
    if columns.get("qty") is None:
        frappe.throw(
            _(
                "QTY column is missing in the BOM import sheet. Add a QTY, QTY., Quantity, or Required Qty header to the quantity column."
            )
        )

    assemblies = []
    main_components = []
    current = None
    rows = []
    for source_row, row in enumerate(table[header_row:], start=header_row + 1):
        row_values = {key: row[index] for key, index in columns.items() if index is not None and index < len(row)}
        if not any(row_values.values()):
            continue

        part_no = _clean_text(row_values.get("part_no"))
        part_name = _clean_text(row_values.get("part_description")) or _clean_text(row_values.get("part_name")) or part_no
        if not part_no:
            part_no, part_name = _use_existing_item_code_from_part_name(part_no, part_name)
        erp_item_code = _clean_text(row_values.get("erp_item_code"))
        qty = flt(row_values.get("qty"))
        uom = _clean_text(row_values.get("uom"))
        row_type = _clean_text(row_values.get("row_type"))
        material = _clean_text(row_values.get("material"))
        bounding_box_length = flt(row_values.get("bounding_box_length"))
        bounding_box_width = flt(row_values.get("bounding_box_width"))
        sheet_metal_thickness = _clean_text(row_values.get("sheet_metal_thickness"))
        mass = flt(row_values.get("mass"))
        gross_weight = flt(row_values.get("gross_weight"))
        rows.append(row_values)

        if _is_main_assembly_row(row_type, row_values):
            main_components.append({
                "source_key": _make_row_source_key(source_row),
                "source_row": source_row,
                "source_part_no": part_no,
                "erp_item_code": erp_item_code,
                "part_name": part_name,
                "qty": qty,
                "uom": uom,
                "row_type": row_type,
                "material": material,
                "bounding_box_length": bounding_box_length,
                "bounding_box_width": bounding_box_width,
                "sheet_metal_thickness": sheet_metal_thickness,
                "mass": mass,
                "gross_weight": gross_weight,
            })
            continue

        if _is_sub_assembly_row(row_type, row_values):
            current = {
                "source_key": _make_row_source_key(source_row),
                "source_row": source_row,
                "source_part_no": part_no,
                "erp_item_code": erp_item_code,
                "part_name": part_name,
                "qty_in_fg": qty,
                "uom": uom,
                "row_type": row_type,
                "material": material,
                "bounding_box_length": bounding_box_length,
                "bounding_box_width": bounding_box_width,
                "sheet_metal_thickness": sheet_metal_thickness,
                "mass": mass,
                "gross_weight": gross_weight,
                "components": [],
            }
            assemblies.append(current)
            continue

        if current:
            current["components"].append({
                "source_key": _make_row_source_key(source_row),
                "source_row": source_row,
                "source_part_no": part_no,
                "erp_item_code": erp_item_code,
                "part_name": part_name,
                "qty": qty,
                "uom": uom,
                "row_type": row_type,
                "material": material,
                "bounding_box_length": bounding_box_length,
                "bounding_box_width": bounding_box_width,
                "sheet_metal_thickness": sheet_metal_thickness,
                "mass": mass,
                "gross_weight": gross_weight,
            })

    return {"assemblies": assemblies, "main_components": main_components, "rows": rows}


def _make_row_source_key(source_row):
    return f"row:{source_row}"


def _use_existing_item_code_from_part_name(part_no, part_name):
    if not part_name:
        return part_no, part_name
    try:
        item_name = frappe.db.get_value("Item", part_name, "item_name")
    except Exception:
        item_name = None
    if item_name:
        return part_name, item_name
    return part_no, part_name


def _detect_header(table):
    best = (0, None, {})
    if not table:
        return None, {}
    for row_number, row in enumerate(table[:25], start=1):
        normalized = [_normalize_header(value) for value in row]
        columns = {key: None for key in HEADER_ALIASES}
        score = 0
        for idx, header in enumerate(normalized):
            for key, aliases in HEADER_ALIASES.items():
                if header in aliases and columns[key] is None:
                    columns[key] = idx
                    score += 1
        if columns["erp_item_code"] is None and columns["part_no"] is not None and columns["part_description"] is not None:
            possible_item_col = columns["part_no"] + 1
            if possible_item_col < columns["part_description"] and not normalized[possible_item_col]:
                columns["erp_item_code"] = possible_item_col
        if columns["qty"] is None:
            columns["qty"] = _infer_qty_column(columns, normalized)
        if score > best[0] and (columns["part_no"] is not None or columns["part_name"] is not None or columns["part_description"] is not None):
            best = (score, row_number, columns)
    return best[1], best[2]


def _infer_qty_column(columns, normalized_headers):
    mass_col = columns.get("mass")
    if mass_col is not None and mass_col + 1 < len(normalized_headers) and not normalized_headers[mass_col + 1]:
        return mass_col + 1

    used_columns = {column for column in columns.values() if column is not None}
    for index in range(len(normalized_headers) - 1, -1, -1):
        if index in used_columns:
            continue
        if not normalized_headers[index]:
            return index
    return None


def _normalize_header(value):
    value = _clean_text(value).lower()
    value = re.sub(r"[_/\-]+", " ", value)
    value = re.sub(r"\s+", " ", value).strip()
    value = value.strip(".")
    return value


def _clean_text(value):
    if value is None:
        return ""
    value = str(value).strip()
    markdown_link = re.match(r"^\[([^\]]+)\]\([^)]+\)$", value)
    if markdown_link:
        value = markdown_link.group(1)
    return value.replace("**", "").strip()


def _is_sub_assembly_row(row_type, row_values):
    if row_type and SUB_ASSEMBLY_RE.search(row_type):
        return True
    # Some exports put the explicit SUB ASSY marker in another mapped column.
    for value in row_values.values():
        value = _clean_text(value)
        if value and SUB_ASSEMBLY_RE.search(value):
            return True
    return False


def _is_main_assembly_row(row_type, row_values):
    if row_type and MAIN_ASSEMBLY_RE.search(row_type):
        return True
    for value in row_values.values():
        value = _clean_text(value)
        if value and MAIN_ASSEMBLY_RE.search(value):
            return True
    return False


def _validate_bom_structure(design_item, parsed):
    errors = []
    fg_item_code = _get_finished_good_item_code(design_item)
    if not design_item.name:
        errors.append(_("Design Request Item is required."))
    if not fg_item_code:
        errors.append(_("Design Request Item must have Final Item Code."))
    elif not frappe.db.exists("Item", fg_item_code):
        errors.append(_("Finished Goods Item {0} does not exist.").format(fg_item_code))
    if not design_item.company:
        errors.append(_("Company is required."))
    if not parsed["assemblies"]:
        errors.append(_("At least one SUB ASSY section is required."))

    seen = {}
    for assembly in parsed["assemblies"]:
        _validate_row_identity(assembly, "Sub-assembly", errors)
        if flt(assembly["qty_in_fg"]) <= 0:
            errors.append(_("Row {0}: sub-assembly quantity must be greater than zero.").format(assembly["source_row"]))
        if not assembly.get("components"):
            errors.append(
                _("Row {0}: sub-assembly {1} has no components. Add component rows below it or remove this SUB ASSY row.").format(
                    assembly["source_row"],
                    assembly.get("source_part_no") or assembly.get("part_name") or "",
                )
            )
        if assembly["source_part_no"] == fg_item_code:
            errors.append(_("Row {0}: FG item cannot be listed as its own child.").format(assembly["source_row"]))
        _track_duplicate_source(assembly, seen, errors)
        for component in assembly["components"]:
            _validate_row_identity(component, "Component", errors)
            if flt(component["qty"]) <= 0:
                errors.append(_("Row {0}: component quantity must be greater than zero.").format(component["source_row"]))
            if component["source_part_no"] and assembly["source_part_no"] and component["source_part_no"] == assembly["source_part_no"]:
                errors.append(_("Row {0}: assembly cannot contain itself.").format(component["source_row"]))
            if component["source_part_no"] == fg_item_code:
                errors.append(_("Row {0}: FG item cannot be listed as its own child.").format(component["source_row"]))
            _track_duplicate_source(component, seen, errors)
    for component in parsed.get("main_components", []):
        _validate_row_identity(component, "Main assembly component", errors)
        if flt(component["qty"]) <= 0:
            errors.append(_("Row {0}: main assembly component quantity must be greater than zero.").format(component["source_row"]))
        if component["source_part_no"] == fg_item_code:
            errors.append(_("Row {0}: FG item cannot be listed as its own child.").format(component["source_row"]))
        _track_duplicate_source(component, seen, errors)

    graph = _build_dependency_graph(parsed)
    cycle = _find_cycle(graph)
    if cycle:
        errors.append(_("Circular assembly dependency detected: {0}").format(" -> ".join(cycle)))

    if errors:
        _set_bom_import_state(design_item.name, "Validation Failed", "\n".join(errors))
        frappe.throw("<br>".join(errors))


def _validate_row_identity(row, label, errors):
    if not row.get("source_part_no") and not row.get("part_name"):
        errors.append(_("{0} row {1}: Part Number or Part Name is required.").format(label, row.get("source_row")))


def _track_duplicate_source(row, seen, errors):
    part_no = row.get("source_part_no")
    if not part_no:
        return
    if _find_existing_item(row):
        return
    signature = (_clean_text(row.get("part_name")).lower(), _clean_text(row.get("uom")).lower())
    if part_no in seen and seen[part_no] != signature:
        error = _("Part Number {0} has conflicting descriptions or UOMs.").format(part_no)
        if error not in errors:
            errors.append(error)
    seen.setdefault(part_no, signature)


def _build_dependency_graph(parsed):
    assembly_by_part_no = {}
    for assembly in parsed["assemblies"]:
        if assembly.get("source_part_no"):
            assembly_by_part_no.setdefault(assembly["source_part_no"], []).append(_get_source_key(assembly))

    graph = {_get_source_key(assembly): set() for assembly in parsed["assemblies"]}
    for assembly in parsed["assemblies"]:
        parent = _get_source_key(assembly)
        if not parent:
            continue
        for component in assembly["components"]:
            matching_assemblies = assembly_by_part_no.get(component.get("source_part_no")) or []
            if len(matching_assemblies) == 1:
                component["assembly_source_key"] = matching_assemblies[0]
                graph[parent].add(matching_assemblies[0])
    return graph


def _find_cycle(graph):
    visited = set()
    stack = []
    active = set()

    def visit(node):
        if node in active:
            return stack[stack.index(node):] + [node]
        if node in visited:
            return None
        active.add(node)
        stack.append(node)
        for child in graph.get(node, []):
            cycle = visit(child)
            if cycle:
                return cycle
        stack.pop()
        active.remove(node)
        visited.add(node)
        return None

    for node in graph:
        cycle = visit(node)
        if cycle:
            return cycle
    return None


def _resolve_or_create_items(design_item, parsed):
    source_to_item = {}
    summary = {"created": [], "reused": [], "barcodes": []}
    created_item_codes = set()
    rows = []
    for assembly in parsed["assemblies"]:
        rows.append((assembly, True))
        rows.extend((component, False) for component in assembly["components"])
    rows.extend((component, False) for component in parsed.get("main_components", []))

    for row, is_assembly in rows:
        key = _get_source_key(row)
        if key in source_to_item:
            continue
        item_code = _find_existing_item(row, created_item_codes)
        if item_code:
            source_to_item[key] = item_code
            summary["reused"].append(item_code)
            continue
        if not is_assembly:
            mapping = _find_mapped_item(row)
            if mapping:
                row["raw_material_item_code"] = mapping.get("erp_item")
                row["raw_material_density"] = flt(mapping.get("material_density"))
        if not frappe.has_permission("Item", "create"):
            frappe.throw(_("You need Create permission on Item to create missing child Items."))
        generated_item_code = _get_next_generated_sub_assembly_code() if is_assembly else _get_next_generated_item_code()
        item_code = _create_missing_item(design_item, row, is_assembly, generated_item_code)
        barcode = _assign_generated_item_barcode(
            item_code,
            preferred_barcode=generated_item_code,
            is_assembly=is_assembly,
        )
        source_to_item[key] = item_code
        summary["created"].append(item_code)
        created_item_codes.add(item_code)
        summary["barcodes"].append(
            {
                "item_code": item_code,
                "barcode": barcode,
                "barcode_type": GENERATED_BARCODE_TYPE,
                "item_name": row.get("part_name") or frappe.db.get_value("Item", item_code, "item_name"),
            }
        )
    return source_to_item, summary


def _get_source_key(row):
    return row.get("source_key") or row.get("source_part_no") or row.get("part_name")


def _find_mapped_item(row):
    description = _normalize_mapping_value(row.get("row_type"))
    thickness = _normalize_thickness(row.get("sheet_metal_thickness"))
    material = _normalize_mapping_value(row.get("material"))
    if not description or not thickness:
        return None

    mappings = frappe.get_all(
        "Design BOM Item Mapping",
        filters={"enabled": 1},
        fields=["sheet_description", "sheet_metal_thickness", "material", "material_density", "erp_item", "priority"],
        order_by="priority desc, modified desc",
    )
    matches = []
    has_description_mapping = False
    for mapping in mappings:
        mapped_description = _normalize_mapping_value(mapping.get("sheet_description"))
        if mapped_description != description:
            continue
        has_description_mapping = True
        if _normalize_thickness(mapping.get("sheet_metal_thickness")) != thickness:
            continue
        mapped_material = _normalize_mapping_value(mapping.get("material"))
        if mapped_material and mapped_material != material:
            continue
        matches.append(mapping)

    if matches:
        matches.sort(
            key=lambda mapping: (
                1 if _normalize_mapping_value(mapping.get("material")) else 0,
                mapping.get("priority") or 0,
            ),
            reverse=True,
        )
        return matches[0]

    if has_description_mapping or _is_sheet_row(row):
        frappe.throw(
            _("Row {0}: No Design BOM Item Mapping found for DESCRIPTION {1}, Sheet Metal Thickness {2}, Material {3}.").format(
                row.get("source_row"),
                row.get("row_type"),
                row.get("sheet_metal_thickness"),
                row.get("material") or "-",
            )
        )
    return None


def _is_sheet_row(row):
    return _normalize_mapping_value(row.get("row_type")) in {"sheet", "sheets"}


def _normalize_mapping_value(value):
    value = re.sub(r"\s+", " ", _clean_text(value)).strip().lower()
    return re.sub(r"\s*#\s*", "#", value)


def _normalize_thickness(value):
    value = _clean_text(value)
    if not value or value == "---":
        return ""
    numeric_value = flt(value)
    if numeric_value:
        return f"{numeric_value:g}"
    return _normalize_mapping_value(value)


def _find_existing_item(row, exclude_item_codes=None):
    exclude_item_codes = set(exclude_item_codes or [])
    erp_item_code = row.get("erp_item_code")
    if erp_item_code and frappe.db.exists("Item", erp_item_code):
        return erp_item_code

    source_part_no = row.get("source_part_no")
    if source_part_no and frappe.db.exists("Item", source_part_no):
        return source_part_no

    engineering_field = _get_engineering_reference_field()
    if engineering_field and source_part_no:
        item_code = frappe.db.get_value("Item", {engineering_field: source_part_no}, "name")
        if item_code:
            return item_code

    return _find_existing_generated_item_by_name(row, exclude_item_codes)


def _find_existing_generated_item_by_name(row, exclude_item_codes=None):
    """Reuse a pre-existing generated Item with the same name, but not one created in this import."""
    exclude_item_codes = set(exclude_item_codes or [])
    item_name = _clean_text(row.get("part_name") or row.get("source_part_no"))
    if not item_name:
        return None

    matches = frappe.get_all(
        "Item",
        filters={"item_name": item_name, "disabled": 0},
        fields=["name", "item_code"],
        order_by="creation asc",
        limit_page_length=20,
    )
    for match in matches:
        item_code = match.get("item_code") or match.get("name")
        if item_code not in exclude_item_codes:
            return match.get("name")
    return None


def _get_engineering_reference_field():
    meta = frappe.get_meta("Item")
    for fieldname in ("custom_part_no", "custom_part_number", "custom_drawing_no", "drawing_no", "drawing_number", "part_number"):
        if meta.has_field(fieldname):
            return fieldname
    return None


def _create_missing_item(design_item, row, is_assembly, generated_item_code=None):
    uom = _get_generated_item_uom(design_item, row, is_assembly)
    if not uom:
        frappe.throw(_("Row {0}: UOM is required to create missing Item.").format(row.get("source_row")))
    if not frappe.db.exists("UOM", uom):
        frappe.throw(_("Row {0}: UOM {1} does not exist.").format(row.get("source_row"), uom))

    item = frappe.new_doc("Item")
    if generated_item_code:
        item.item_code = generated_item_code
    item.item_name = row.get("part_name") or row.get("source_part_no")
    item.description = row.get("part_name") or row.get("source_part_no")
    item.item_group = _get_default_item_group(design_item, is_assembly)
    item.stock_uom = uom
    item.is_stock_item = 1
    if _is_sheet_row(row) and frappe.get_meta("Item").has_field("is_sub_contracted_item"):
        item.is_sub_contracted_item = 1
    if frappe.get_meta("Item").has_field("gst_hsn_code"):
        hsn_code = frappe.db.get_value("Item", _get_finished_good_item_code(design_item), "gst_hsn_code")
        if hsn_code:
            item.gst_hsn_code = hsn_code

    engineering_field = _get_engineering_reference_field()
    if engineering_field and row.get("source_part_no") and row.get("source_part_no") != item.item_name:
        item.set(engineering_field, row.get("source_part_no"))

    item.insert()
    desired_item_code = _clean_text(generated_item_code or row.get("source_part_no"))
    if desired_item_code and item.name != desired_item_code and not frappe.db.exists("Item", desired_item_code):
        frappe.rename_doc("Item", item.name, desired_item_code, force=True)
        item.name = desired_item_code
        frappe.db.set_value("Item", desired_item_code, "item_code", desired_item_code, update_modified=False)
    return item.name


def _assign_generated_item_barcode(item_code, preferred_barcode=None, is_assembly=False):
    existing_barcode = frappe.db.get_value(
        "Item Barcode",
        {"parent": item_code, "barcode_type": GENERATED_BARCODE_TYPE},
        "barcode",
    )
    if existing_barcode:
        return existing_barcode

    item = frappe.get_doc("Item", item_code)
    attempted = set()
    for _attempt in range(100):
        if is_assembly:
            barcode = preferred_barcode if preferred_barcode and preferred_barcode not in attempted else _get_next_generated_sub_assembly_barcode()
        else:
            barcode = preferred_barcode if preferred_barcode and preferred_barcode not in attempted else _get_next_generated_barcode()
        attempted.add(barcode)
        if frappe.db.exists("Item Barcode", {"barcode": barcode}):
            continue

        item.append(
            "barcodes",
            {
                "barcode": barcode,
                "barcode_type": GENERATED_BARCODE_TYPE,
                "uom": item.stock_uom,
            },
        )
        try:
            item.save(ignore_permissions=True)
            return barcode
        except (frappe.DuplicateEntryError, frappe.UniqueValidationError):
            item.set("barcodes", [row for row in item.get("barcodes") if row.get("barcode") != barcode])
            item.reload()

    if is_assembly:
        frappe.throw(_("Could not generate a unique SUB barcode for Item {0}.").format(item_code))
    frappe.throw(_("Could not generate a unique 6 digit barcode for Item {0}.").format(item_code))


def _get_next_generated_item_code():
    for _attempt in range(100):
        item_code = _get_next_generated_barcode()
        if frappe.db.exists("Item", item_code) or frappe.db.exists("Item Barcode", {"barcode": item_code}):
            continue
        return item_code
    frappe.throw(_("Could not generate a unique 6 digit Item Code."))


def _get_next_generated_sub_assembly_code():
    for _attempt in range(100):
        item_code = _get_next_generated_sub_assembly_barcode()
        if frappe.db.exists("Item", item_code) or frappe.db.exists("Item Barcode", {"barcode": item_code}):
            continue
        return item_code
    frappe.throw(_("Could not generate a unique SUB Item Code."))


def _get_next_generated_barcode():
    result = frappe.db.sql(
        """
        select code
        from (
            select barcode as code
            from `tabItem Barcode`
            where barcode regexp '^[0-9]{6}$'
            union
            select item_code as code
            from `tabItem`
            where item_code regexp '^[0-9]{6}$'
        ) generated_codes
        order by code desc
        limit 1
        """
    )
    next_number = (int(result[0][0]) + 1) if result else 1
    if next_number > 999999:
        frappe.throw(_("No 6 digit barcode numbers are available."))
    return f"{next_number:06d}"


def _get_next_generated_sub_assembly_barcode():
    result = frappe.db.sql(
        """
        select code
        from (
            select barcode as code
            from `tabItem Barcode`
            where barcode regexp '^SUB[0-9]{5}$'
            union
            select item_code as code
            from `tabItem`
            where item_code regexp '^SUB[0-9]{5}$'
        ) generated_codes
        order by code desc
        limit 1
        """
    )
    next_number = (int(result[0][0][3:]) + 1) if result else 1
    if next_number > 99999:
        frappe.throw(_("No SUB barcode numbers are available."))
    return f"SUB{next_number:05d}"


def _get_generated_item_uom(design_item, row, is_assembly):
    if row.get("raw_material_item_code") and frappe.db.exists("UOM", "Nos"):
        return "Nos"
    return row.get("uom") or design_item.uom


def _get_default_item_group(design_item, is_assembly):
    fg_item_code = _get_finished_good_item_code(design_item)
    if fg_item_code:
        item_group = frappe.db.get_value("Item", fg_item_code, "item_group")
        if item_group:
            return item_group
    if frappe.db.exists("Item Group", "Products"):
        return "Products"
    frappe.throw(_("No default Item Group found for generated Items."))


def _create_bom_hierarchy(design_item, parsed, source_to_item):
    graph = _build_dependency_graph(parsed)
    assembly_map = {_get_source_key(assembly): assembly for assembly in parsed["assemblies"]}
    assembly_boms = {}
    component_boms = {}
    summary = {"created": [], "reused": []}

    def make_component_row(component):
        component_key = _get_source_key(component)
        component_item = source_to_item[component_key]
        child_bom = assembly_boms.get(component.get("assembly_source_key")) or component_boms.get(component_key) or _get_default_bom(component_item, design_item.company)
        if component.get("raw_material_item_code") and not child_bom:
            child_bom, reused = _create_sheet_component_bom(design_item, component_item, component)
            component_boms[component_key] = child_bom
            summary["reused" if reused else "created"].append(child_bom)
        return _make_bom_row(component_item, flt(component["qty"]), None if child_bom else component.get("uom"), child_bom, component.get("source_row"))

    def build_for(source):
        if source in assembly_boms:
            return assembly_boms[source]
        for child in graph.get(source, []):
            build_for(child)

        assembly = assembly_map[source]
        item_code = source_to_item[source]

        rows = []
        for component in assembly["components"]:
            rows.append(make_component_row(component))
        rows = _combine_bom_rows(rows)

        bom_name, reused = _get_or_create_submitted_bom(
            item_code=item_code,
            company=design_item.company,
            quantity=1,
            rows=rows,
            is_default=0,
        )
        assembly_boms[source] = bom_name
        summary["reused" if reused else "created"].append(bom_name)
        return bom_name

    child_sources = {child for children in graph.values() for child in children}
    top_level_sources = [_get_source_key(assembly) for assembly in parsed["assemblies"] if _get_source_key(assembly) not in child_sources]
    for source in top_level_sources:
        build_for(source)

    fg_rows = []
    for source in top_level_sources:
        assembly = assembly_map[source]
        item_code = source_to_item[source]
        fg_rows.append(_make_bom_row(item_code, assembly["qty_in_fg"], assembly.get("uom"), assembly_boms[source], assembly.get("source_row")))
    for component in parsed.get("main_components", []):
        fg_rows.append(make_component_row(component))
    fg_rows = _combine_bom_rows(fg_rows)

    fg_bom, reused = _get_or_create_submitted_bom(
        item_code=_get_finished_good_item_code(design_item),
        company=design_item.company,
        quantity=1,
        rows=fg_rows,
        is_default=1,
    )
    assembly_boms["__fg__"] = fg_bom
    summary["reused" if reused else "created"].append(fg_bom)
    return assembly_boms, summary


def _create_sheet_component_bom(design_item, item_code, component):
    raw_material_item = component.get("raw_material_item_code")
    if not raw_material_item:
        return None, True
    raw_qty = _get_sheet_raw_material_qty(component)
    if raw_qty <= 0:
        frappe.throw(_("Row {0}: Raw material quantity could not be calculated for sheet BOM.").format(component.get("source_row")))
    raw_stock_uom = frappe.db.get_value("Item", raw_material_item, "stock_uom")
    rows = [_make_bom_row(raw_material_item, raw_qty, raw_stock_uom, None, component.get("source_row"))]
    return _get_or_create_submitted_bom(
        item_code=item_code,
        company=design_item.company,
        quantity=1,
        rows=rows,
        is_default=1,
        scrap_rows=[{"item_code": raw_material_item, "stock_qty": 1}],
    )


def _get_sheet_raw_material_qty(component):
    density = flt(component.get("raw_material_density"))
    length = flt(component.get("bounding_box_length"))
    width = flt(component.get("bounding_box_width"))
    thickness = flt(component.get("sheet_metal_thickness"))
    if density > 0 and length > 0 and width > 0 and thickness > 0:
        return length * width * thickness * density / 1000000
    return flt(component.get("gross_weight")) or flt(component.get("mass"))


@frappe.whitelist()
def repair_generated_sheet_bom_raw_material_quantities(design_request_item=None, root_bom=None, bom_file=None, dry_run=1, use_history=0):
    """Repair generated sheet-part BOM raw-material qty from the original imported sheet."""
    dry_run = cint(dry_run)
    if cint(use_history):
        return _repair_generated_sheet_bom_raw_material_quantities_from_history(design_request_item, dry_run)
    if root_bom:
        return _repair_generated_sheet_bom_raw_material_quantities_for_root_bom(root_bom, bom_file, dry_run)

    design_items = _get_design_items_for_sheet_bom_repair(design_request_item)
    repairs = []
    rebuilt_boms = set()

    for design_item in design_items:
        try:
            fg_item_code = _get_finished_good_item_code(design_item)
            workbook_path = _resolve_bom_workbook(design_item)
            parsed = _parse_bom_workbook(workbook_path, fg_item_code)
            _apply_mappings_to_sheet_components(parsed)
            source_to_item = _resolve_existing_items_for_repair(parsed)
            assembly_boms = _get_existing_bom_hierarchy_for_repair(design_item, parsed, source_to_item)
            component_boms = _repair_sheet_component_boms_from_parsed_sheet(
                design_item,
                parsed,
                source_to_item,
                dry_run,
                repairs,
            )
            if not dry_run and component_boms:
                for bom_name in _get_repair_rebuild_order(assembly_boms.values(), component_boms):
                    _rebuild_bom_totals(bom_name)
                    rebuilt_boms.add(bom_name)
        except Exception:
            frappe.log_error(frappe.get_traceback(), "Design Sheet BOM Quantity Repair Failed")
            repairs.append({
                "design_request_item": design_item.name,
                "error": frappe.get_traceback(),
            })

    if not dry_run:
        frappe.db.commit()

    return {
        "dry_run": dry_run,
        "design_item_count": len(design_items),
        "repair_count": len([row for row in repairs if not row.get("error")]),
        "rebuilt_bom_count": len(rebuilt_boms),
        "repairs": repairs,
    }


def _repair_generated_sheet_bom_raw_material_quantities_from_history(design_request_item, dry_run):
    if not design_request_item:
        frappe.throw(_("Design Request Item is required when repairing from history."))

    targets = _get_sheet_bom_repair_targets_from_history(design_request_item)
    repairs = []
    errors = []
    rebuilt_bom_count = 0

    for target in targets:
        try:
            result = _repair_generated_sheet_bom_raw_material_quantities_for_root_bom(
                target.get("root_bom"),
                target.get("bom_file"),
                dry_run,
                commit=False,
            )
            repairs.extend(result.get("repairs") or [])
            rebuilt_bom_count += cint(result.get("rebuilt_bom_count"))
        except Exception:
            frappe.log_error(frappe.get_traceback(), "Design Sheet BOM History Repair Failed")
            errors.append({
                "root_bom": target.get("root_bom"),
                "bom_file": target.get("bom_file"),
                "error": frappe.get_traceback(),
            })

    if not dry_run:
        frappe.db.commit()

    return {
        "dry_run": dry_run,
        "design_request_item": design_request_item,
        "target_count": len(targets),
        "repair_count": len([row for row in repairs if not row.get("error")]),
        "rebuilt_bom_count": rebuilt_bom_count,
        "targets": targets,
        "repairs": repairs,
        "errors": errors,
    }


def _get_sheet_bom_repair_targets_from_history(design_request_item):
    versions = frappe.get_all(
        "Version",
        filters={"ref_doctype": "Design Request Item", "docname": design_request_item},
        fields=["name", "creation", "data"],
        order_by="creation asc",
        limit=1000,
    )
    targets = []
    seen = set()
    current_file = None
    last_file = None
    current_item = None

    for version in versions:
        data = frappe.parse_json(version.data or "{}")
        changed = data.get("changed") or []
        bom_cleared = None
        file_cleared = None
        item_hint = None

        for row in changed:
            if not row or len(row) < 3:
                continue
            fieldname, old_value, new_value = row[0], row[1], row[2]
            if fieldname == "new_item_code":
                if new_value:
                    current_item = new_value
                elif old_value:
                    item_hint = old_value
            elif fieldname in ("custom_bom_for_import", "custom_bom_importer"):
                if new_value:
                    current_file = new_value
                    last_file = new_value
                elif old_value:
                    file_cleared = old_value
                    last_file = old_value
                    current_file = None
            elif fieldname == "bom_name" and old_value and not new_value:
                bom_cleared = old_value

        if not bom_cleared:
            continue

        bom_file = file_cleared or current_file or last_file
        if not bom_file:
            continue

        key = (bom_cleared, bom_file)
        if key in seen:
            continue
        seen.add(key)
        targets.append({
            "root_bom": bom_cleared,
            "bom_file": bom_file,
            "version": version.name,
            "version_creation": version.creation,
            "item_code": item_hint or current_item,
        })

    return targets


def _repair_generated_sheet_bom_raw_material_quantities_for_root_bom(root_bom, bom_file, dry_run, commit=True):
    if not bom_file:
        frappe.throw(_("BOM file is required when repairing by root BOM."))
    if not frappe.db.exists("BOM", root_bom):
        frappe.throw(_("BOM {0} was not found.").format(root_bom))

    root_bom_doc = frappe.get_doc("BOM", root_bom)
    design_item = frappe._dict({
        "name": root_bom,
        "item_code": root_bom_doc.item,
        "new_item_code": root_bom_doc.item,
        "company": root_bom_doc.company,
        "bom_name": root_bom,
        "custom_bom_importer": bom_file,
    })
    workbook_path = _resolve_bom_workbook(design_item)
    parsed = _parse_bom_workbook(workbook_path, root_bom_doc.item)
    _apply_mappings_to_sheet_components(parsed)

    repairs = []
    component_boms = _repair_sheet_component_boms_by_bom_tree_paths(
        root_bom,
        parsed,
        dry_run,
        repairs,
    )
    rebuilt_boms = set()
    if not dry_run and component_boms:
        for bom_name in _get_repair_rebuild_order([root_bom], component_boms):
            _rebuild_bom_totals(bom_name)
            rebuilt_boms.add(bom_name)
        if commit:
            frappe.db.commit()

    return {
        "dry_run": dry_run,
        "root_bom": root_bom,
        "bom_file": bom_file,
        "design_item_count": 0,
        "repair_count": len([row for row in repairs if not row.get("error")]),
        "rebuilt_bom_count": len(rebuilt_boms),
        "repairs": repairs,
    }


def _get_design_items_for_sheet_bom_repair(design_request_item=None):
    if design_request_item:
        return [frappe.get_doc("Design Request Item", design_request_item)]

    filters = {"bom_created": 1, "bom_name": ["is", "set"]}
    meta = frappe.get_meta("Design Request Item")
    if meta.has_field("custom_bom_import_status"):
        filters["custom_bom_import_status"] = "Completed"

    return frappe.get_all("Design Request Item", filters=filters, fields="*", order_by="modified asc")


def _apply_mappings_to_sheet_components(parsed):
    for component in _iter_parsed_components(parsed):
        if not _is_sheet_row(component):
            continue
        mapping = _find_mapped_item(component)
        component["raw_material_item_code"] = mapping.get("erp_item")
        component["raw_material_density"] = flt(mapping.get("material_density"))


def _iter_parsed_components(parsed):
    for assembly in parsed.get("assemblies", []):
        for component in assembly.get("components", []):
            yield component
    for component in parsed.get("main_components", []):
        yield component


def _resolve_existing_items_for_repair(parsed):
    source_to_item = {}
    for assembly in parsed.get("assemblies", []):
        source_to_item[_get_source_key(assembly)] = _find_existing_item(assembly)
        for component in assembly.get("components", []):
            source_to_item[_get_source_key(component)] = _find_existing_item(component)
    for component in parsed.get("main_components", []):
        source_to_item[_get_source_key(component)] = _find_existing_item(component)
    return {key: item for key, item in source_to_item.items() if item}


def _get_existing_bom_hierarchy_for_repair(design_item, parsed, source_to_item):
    graph = _build_dependency_graph(parsed)
    assembly_boms = {}

    def build_for(source):
        if source in assembly_boms:
            return assembly_boms[source]
        for child in graph.get(source, []):
            build_for(child)
        item_code = source_to_item.get(source)
        if item_code:
            assembly_boms[source] = _get_default_bom(item_code, design_item.company)
        return assembly_boms.get(source)

    child_sources = {child for children in graph.values() for child in children}
    top_level_sources = [_get_source_key(assembly) for assembly in parsed["assemblies"] if _get_source_key(assembly) not in child_sources]
    for source in top_level_sources:
        build_for(source)

    fg_bom = getattr(design_item, "bom_name", None) or _get_default_bom(_get_finished_good_item_code(design_item), design_item.company)
    if fg_bom:
        assembly_boms["__fg__"] = fg_bom

    return {source: bom for source, bom in assembly_boms.items() if bom}


def _repair_sheet_component_boms_from_parsed_sheet(design_item, parsed, source_to_item, dry_run, repairs):
    component_boms = set()

    for component in _iter_parsed_components(parsed):
        raw_material_item = component.get("raw_material_item_code")
        if not raw_material_item:
            continue
        component_item = source_to_item.get(_get_source_key(component))
        if not component_item:
            continue
        bom_name = _get_default_bom(component_item, design_item.company)
        if not bom_name:
            continue
        bom = frappe.get_doc("BOM", bom_name)
        if not _is_single_raw_material_bom(bom):
            continue

        raw_row = bom.items[0]
        expected_qty = _get_sheet_raw_material_qty(component)
        current_qty = flt(raw_row.stock_qty)
        if raw_row.item_code != raw_material_item or expected_qty <= 0 or abs(current_qty - expected_qty) <= 0.000001:
            continue

        repairs.append({
            "design_request_item": design_item.name,
            "bom": bom.name,
            "item": bom.item,
            "raw_material_item": raw_material_item,
            "source_row": component.get("source_row"),
            "current_qty": current_qty,
            "expected_qty": expected_qty,
        })
        component_boms.add(bom.name)

        if not dry_run:
            _update_bom_raw_material_row(raw_row, expected_qty)

    return component_boms


def _repair_sheet_component_boms_by_bom_tree_paths(root_bom, parsed, dry_run, repairs):
    parsed_components = _get_sheet_components_by_repair_path(parsed)
    bom_components = _get_single_raw_material_boms_by_repair_path(root_bom)
    component_boms = set()

    for path, components in parsed_components.items():
        bom_names = bom_components.get(path) or []
        for index, component in enumerate(components):
            if index >= len(bom_names):
                continue
            bom = frappe.get_doc("BOM", bom_names[index])
            raw_row = bom.items[0]
            raw_material_item = component.get("raw_material_item_code")
            expected_qty = _get_sheet_raw_material_qty(component)
            current_qty = flt(raw_row.stock_qty)
            if raw_row.item_code != raw_material_item or expected_qty <= 0 or abs(current_qty - expected_qty) <= 0.000001:
                continue

            repairs.append({
                "root_bom": root_bom,
                "path": path,
                "bom": bom.name,
                "item": bom.item,
                "raw_material_item": raw_material_item,
                "source_row": component.get("source_row"),
                "current_qty": current_qty,
                "expected_qty": expected_qty,
            })
            component_boms.add(bom.name)

            if not dry_run:
                _update_bom_raw_material_row(raw_row, expected_qty)

    return component_boms


def _get_sheet_components_by_repair_path(parsed):
    graph = _build_dependency_graph(parsed)
    assembly_map = {_get_source_key(assembly): assembly for assembly in parsed.get("assemblies", [])}
    child_sources = {child for children in graph.values() for child in children}
    top_level_sources = [_get_source_key(assembly) for assembly in parsed.get("assemblies", []) if _get_source_key(assembly) not in child_sources]
    components_by_path = {}

    def walk_assembly(source, parent_path):
        assembly = assembly_map[source]
        assembly_path = parent_path + [_get_repair_path_label(assembly)]
        for component in assembly.get("components", []):
            component_path = assembly_path + [_get_repair_path_label(component)]
            child_source = component.get("assembly_source_key")
            if child_source and child_source in assembly_map:
                walk_assembly(child_source, assembly_path)
            elif component.get("raw_material_item_code"):
                _append_repair_path_row(components_by_path, component_path, component)

    for source in top_level_sources:
        walk_assembly(source, [])
    for component in parsed.get("main_components", []):
        if component.get("raw_material_item_code"):
            _append_repair_path_row(components_by_path, [_get_repair_path_label(component)], component)
    return components_by_path


def _get_single_raw_material_boms_by_repair_path(root_bom):
    boms_by_path = {}

    def walk_bom(bom_name, parent_path):
        bom = frappe.get_doc("BOM", bom_name)
        for row in bom.items:
            row_path = parent_path + [_get_repair_path_label(row)]
            if row.bom_no:
                child_bom = frappe.get_doc("BOM", row.bom_no)
                if _is_single_raw_material_bom(child_bom):
                    _append_repair_path_row(boms_by_path, row_path, child_bom.name)
                else:
                    walk_bom(child_bom.name, row_path)

    walk_bom(root_bom, [])
    return boms_by_path


def _append_repair_path_row(rows_by_path, path_parts, row):
    key = " > ".join(path_parts)
    rows_by_path.setdefault(key, []).append(row)


def _get_repair_path_label(row):
    get = row.get if hasattr(row, "get") else lambda key: getattr(row, key, None)
    value = get("item_name") or get("description") or get("part_name") or get("source_part_no") or get("item_code")
    return re.sub(r"\s+", " ", _clean_text(value)).strip().upper()


def _is_single_raw_material_bom(bom):
    return bom.docstatus == 1 and len(bom.items) == 1 and not bom.items[0].bom_no


def _update_bom_raw_material_row(row, stock_qty):
    stock_qty = flt(stock_qty)
    conversion_factor = flt(row.conversion_factor) or 1
    rate = flt(row.rate)
    base_rate = flt(row.base_rate) or rate
    frappe.db.set_value(
        "BOM Item",
        row.name,
        {
            "qty": stock_qty / conversion_factor,
            "stock_qty": stock_qty,
            "qty_consumed_per_unit": stock_qty,
            "amount": stock_qty * rate,
            "base_amount": stock_qty * base_rate,
        },
        update_modified=False,
    )


def _get_repair_rebuild_order(assembly_boms, component_boms):
    ordered = []
    seen = set()
    for bom_name in list(component_boms) + list(assembly_boms):
        if not bom_name:
            continue
        for tree_bom in _get_bom_tree_postorder(bom_name):
            if tree_bom not in seen:
                seen.add(tree_bom)
                ordered.append(tree_bom)
    return ordered


def _get_default_bom(item_code, company=None):
    default_bom = frappe.db.get_value("Item", item_code, "default_bom")
    if default_bom and frappe.db.get_value("BOM", default_bom, "docstatus") == 1:
        if not company or frappe.db.get_value("BOM", default_bom, "company") == company:
            return default_bom
    return frappe.db.get_value("BOM", {"item": item_code, "company": company, "is_default": 1, "docstatus": 1}, "name")


@frappe.whitelist()
def repair_bom_exploded_items(root_bom=None, dry_run=1, only_generated=1, repeat_until_clean=0):
    """Rebuild stored BOM Explosion Items when the visible BOM tree is correct."""
    dry_run = cint(dry_run)
    only_generated = cint(only_generated)
    repeat_until_clean = cint(repeat_until_clean)
    bom_names = _get_boms_for_exploded_repair(root_bom, only_generated)
    repairs = []
    max_passes = 3 if repeat_until_clean and not dry_run else 1

    for pass_number in range(1, max_passes + 1):
        pass_repairs = []
        for bom_name in bom_names:
            bom = frappe.get_doc("BOM", bom_name)
            expected_signature = _expected_exploded_signature(bom)
            current_signature = _exploded_signature(bom.exploded_items)
            if expected_signature == current_signature:
                continue

            repair = {
                "bom": bom.name,
                "item": bom.item,
                "pass": pass_number,
                "current_rows": len(current_signature),
                "expected_rows": len(expected_signature),
            }
            pass_repairs.append(repair)
            repairs.append(repair)
            if not dry_run:
                _rebuild_bom_totals(bom.name)

        if not pass_repairs:
            break

    if not dry_run:
        frappe.db.commit()

    return {
        "dry_run": dry_run,
        "root_bom": root_bom,
        "only_generated": only_generated,
        "repeat_until_clean": repeat_until_clean,
        "bom_count": len(bom_names),
        "repair_count": len(repairs),
        "repairs": repairs,
    }


def _get_boms_for_exploded_repair(root_bom=None, only_generated=1):
    if root_bom:
        if not frappe.db.exists("BOM", root_bom):
            frappe.throw(_("BOM {0} was not found.").format(root_bom))
        return _get_bom_tree_postorder(root_bom)

    boms = frappe.get_all("BOM", filters={"docstatus": 1}, fields=["name", "item"], order_by="modified asc")
    names = [
        bom.name
        for bom in boms
        if not only_generated or _is_generated_bom_item(bom.item)
    ]
    ordered = []
    seen = set()
    for name in names:
        for bom_name in _get_bom_tree_postorder(name):
            if bom_name not in seen:
                seen.add(bom_name)
                ordered.append(bom_name)
    return ordered


def _is_generated_bom_item(item_code):
    item_code = _clean_text(item_code)
    return bool(re.fullmatch(r"\d{6}", item_code) or re.fullmatch(r"SUB\d{5}", item_code) or item_code.startswith("P-SY-CE-"))


def _get_bom_tree_postorder(root_bom):
    visited = set()
    ordered = []

    def walk(bom_name):
        if bom_name in visited:
            return
        visited.add(bom_name)
        for child_bom in frappe.get_all("BOM Item", filters={"parent": bom_name, "bom_no": ["is", "set"]}, pluck="bom_no"):
            walk(child_bom)
        ordered.append(bom_name)

    walk(root_bom)
    return ordered


def _expected_exploded_signature(bom):
    return _exploded_signature(_get_expected_exploded_rows_from_bom(bom.name))


def _get_expected_exploded_rows_from_bom(bom_name, multiplier=1, visiting=None):
    visiting = visiting or set()
    if bom_name in visiting:
        frappe.throw(_("Circular BOM dependency detected while repairing exploded items: {0}").format(bom_name))

    visiting.add(bom_name)
    bom = frappe.get_doc("BOM", bom_name)
    rows = {}

    for item in bom.items:
        if item.bom_no:
            child_bom_qty = flt(frappe.db.get_value("BOM", item.bom_no, "quantity")) or 1
            child_multiplier = multiplier * flt(item.stock_qty) / child_bom_qty
            for child_row in _get_expected_exploded_rows_from_bom(item.bom_no, child_multiplier, visiting):
                _add_expected_exploded_row(rows, child_row)
        elif item.item_code:
            _add_expected_exploded_row(rows, {
                "item_code": item.item_code,
                "stock_qty": multiplier * flt(item.stock_qty),
                "stock_uom": item.stock_uom,
                "include_item_in_manufacturing": item.include_item_in_manufacturing,
                "sourced_by_supplier": item.sourced_by_supplier,
            })

    visiting.remove(bom_name)
    return rows.values()


def _add_expected_exploded_row(rows, row):
    item_code = row.get("item_code")
    if not item_code:
        return

    if item_code in rows:
        rows[item_code]["stock_qty"] += flt(row.get("stock_qty"))
        return

    rows[item_code] = frappe._dict({
        "item_code": item_code,
        "stock_qty": flt(row.get("stock_qty")),
        "stock_uom": row.get("stock_uom"),
        "include_item_in_manufacturing": cint(row.get("include_item_in_manufacturing")),
        "sourced_by_supplier": cint(row.get("sourced_by_supplier")),
    })


def _exploded_signature(rows):
    signature = []
    for row in rows:
        get = row.get if hasattr(row, "get") else lambda key: getattr(row, key, None)
        signature.append((
            get("item_code"),
            flt(get("stock_qty"), 6),
            get("stock_uom"),
            cint(get("include_item_in_manufacturing")),
            cint(get("sourced_by_supplier")),
        ))
    return sorted(signature)


def _rebuild_bom_totals(bom_name):
    bom = frappe.get_doc("BOM", bom_name)
    bom.update_exploded_items(save=True)
    bom.calculate_rm_cost(save=True)
    bom.calculate_sm_cost(save=True)
    bom.calculate_exploded_cost()
    bom.db_update()


def _make_bom_row(item_code, qty, source_uom, child_bom, source_row):
    stock_uom = frappe.db.get_value("Item", item_code, "stock_uom")
    conversion_factor = _get_conversion_factor(item_code, source_uom, stock_uom, source_row)
    return {
        "item_code": item_code,
        "qty": flt(qty),
        "uom": source_uom or stock_uom,
        "stock_uom": stock_uom,
        "conversion_factor": conversion_factor,
        "bom_no": child_bom,
    }


def _combine_bom_rows(rows):
    combined = {}
    for row in rows:
        key = (
            row.get("item_code"),
            row.get("uom"),
            row.get("stock_uom"),
            flt(row.get("conversion_factor")),
            row.get("bom_no") or None,
        )
        if key not in combined:
            combined[key] = dict(row)
            continue
        combined[key]["qty"] = flt(combined[key].get("qty")) + flt(row.get("qty"))
    return list(combined.values())


def _get_conversion_factor(item_code, source_uom, stock_uom, source_row):
    if not source_uom or source_uom == stock_uom:
        return 1
    factor = frappe.db.get_value("UOM Conversion Detail", {"parent": item_code, "uom": source_uom}, "conversion_factor")
    if not factor:
        frappe.throw(_("Row {0}: UOM {1} does not match Stock UOM {2} for Item {3}, and no conversion exists.").format(source_row, source_uom, stock_uom, item_code))
    return flt(factor)


def _get_or_create_submitted_bom(item_code, company, quantity, rows, is_default=0, scrap_rows=None):
    scrap_rows = scrap_rows or []
    existing_boms = frappe.get_all(
        "BOM",
        filters={"item": item_code, "company": company, "docstatus": ["in", [0, 1]]},
        fields=["name", "docstatus"],
        order_by="docstatus desc, modified desc",
    )
    target_signature = _bom_signature(rows)
    target_scrap_signature = _bom_scrap_signature(scrap_rows)
    for existing in existing_boms:
        bom = frappe.get_doc("BOM", existing.name)
        if _bom_signature(bom.items) == target_signature and _bom_scrap_signature(bom.scrap_items) == target_scrap_signature:
            if bom.docstatus == 0:
                bom.submit()
            return bom.name, True

    bom = frappe.new_doc("BOM")
    bom.item = item_code
    bom.company = company
    bom.quantity = quantity
    bom.is_active = 1
    bom.is_default = is_default
    for row in rows:
        bom.append("items", row)
    for row in scrap_rows:
        bom.append("scrap_items", row)
    bom.insert()
    bom.submit()
    if is_default:
        frappe.db.set_value("Item", item_code, "default_bom", bom.name, update_modified=False)
    return bom.name, False


def _bom_scrap_signature(rows):
    signature = []
    for row in rows:
        get = row.get if hasattr(row, "get") else lambda key: getattr(row, key, None)
        signature.append((
            get("item_code"),
            flt(get("stock_qty")),
        ))
    return sorted(signature)


def _bom_signature(rows):
    signature = []
    for row in rows:
        get = row.get if hasattr(row, "get") else lambda key: getattr(row, key, None)
        signature.append((
            get("item_code"),
            flt(get("qty")),
            get("uom"),
            get("bom_no") or None,
        ))
    return sorted(signature)
